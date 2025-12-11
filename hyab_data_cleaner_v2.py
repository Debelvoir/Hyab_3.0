"""
HYAB Data Cleaner + Analytics + Intelligence v2.1
Updated based on Victor's feedback (Dec 2025)

Changes in v2.1:
- Fixed sheet name detection (accepts "Sheet1", "Order book", etc.)
- Sales: Now processes both Article and Company sheets
- Sales: Uses correct quantity column (Antal ut)
- Sales: Added Top 20 Articles/Customers analysis with charts
- Sales: Flags TB/TG as potentially unreliable
- Order Book: Added delivery date placeholder column
"""

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime, timedelta
from io import BytesIO
from collections import defaultdict
import plotly.graph_objects as go
import plotly.express as px

# --- Page Config ---
st.set_page_config(
    page_title="HYAB Data Cleaner",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# --- Inject Custom Styles ---
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Libre+Baskerville:ital@1&family=Source+Sans+3:wght@400;500;600&display=swap');

:root {
    --navy: #1B2A41;
    --navy-dark: #141F30;
    --grey-warm: #F7F6F4;
    --grey-border: #E2E0DC;
    --grey-text: #6B7280;
    --teal: #4A9BA8;
    --success: #2D6A4F;
    --success-light: #D8F3DC;
    --warning: #9A6700;
    --warning-light: #FEF3CD;
    --danger: #9B2C2C;
    --danger-light: #FED7D7;
}

#MainMenu, footer, header, .stDeployButton {display: none !important;}
section[data-testid="stSidebar"] {display: none;}
.stApp {background: var(--grey-warm);}
.block-container {max-width: 720px !important; padding: 0 1rem 4rem 1rem !important;}

.hyab-header {
    background: var(--navy);
    padding: 14px 24px;
    margin: -1rem -1rem 1.5rem -1rem;
    display: flex;
    align-items: center;
    gap: 12px;
    font-family: 'Source Sans 3', sans-serif;
}
.hyab-header .brand-mark {color: var(--teal); font-size: 22px;}
.hyab-header .brand-text {color: white; font-size: 13px; font-weight: 600; letter-spacing: 0.12em;}
.hyab-header .brand-divider {color: rgba(255,255,255,0.25); margin: 0 4px;}
.hyab-header .brand-page {color: rgba(255,255,255,0.7); font-size: 13px;}

.page-title {
    font-family: 'Libre Baskerville', Georgia, serif;
    font-size: 26px;
    font-style: italic;
    font-weight: 400;
    color: var(--navy);
    margin: 0 0 20px 0;
}

.section-label {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--grey-text);
    margin: 0 0 10px 0;
}

div[data-testid="stRadio"] > div {
    flex-direction: row !important;
    gap: 0 !important;
    background: white;
    border: 1px solid var(--grey-border);
    padding: 3px;
    width: fit-content;
}
div[data-testid="stRadio"] > div > label {
    padding: 8px 18px !important;
    margin: 0 !important;
    border-radius: 0 !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    background: transparent !important;
}
div[data-testid="stRadio"] > div > label[data-checked="true"] {
    background: var(--navy) !important;
    color: white !important;
}
div[data-testid="stRadio"] > label {display: none !important;}

input[type="number"], .stNumberInput input {
    border: 1px solid var(--grey-border) !important;
    border-radius: 0 !important;
    font-size: 15px !important;
    height: 42px !important;
    text-align: right !important;
}
input:focus {border-color: var(--navy) !important; box-shadow: none !important;}

section[data-testid="stFileUploader"] {
    border: 2px dashed var(--grey-border) !important;
    border-radius: 0 !important;
    background: white !important;
}
section[data-testid="stFileUploader"]:hover {border-color: var(--navy) !important;}

.stButton > button {
    background: var(--navy) !important;
    color: white !important;
    border: none !important;
    border-radius: 0 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    height: 48px !important;
    width: 100% !important;
}
.stButton > button:hover {background: var(--navy-dark) !important;}
.stButton > button:disabled {background: var(--grey-border) !important; color: var(--grey-text) !important;}

.stDownloadButton > button {
    background: var(--success) !important;
    color: white !important;
    border: none !important;
    border-radius: 0 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    height: 48px !important;
}
.stDownloadButton > button:hover {background: #245840 !important;}

div[data-testid="stMetric"] {
    background: var(--grey-warm);
    padding: 16px;
    border-left: 3px solid var(--navy);
}
div[data-testid="stMetric"] label {
    font-size: 11px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
    color: var(--grey-text) !important;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    font-size: 24px !important;
    font-weight: 600 !important;
    color: var(--navy) !important;
}

.alert-grid {display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; margin: 12px 0;}
.alert-box {padding: 14px 10px; font-size: 12px; font-weight: 500; text-align: center;}
.alert-box .icon {font-size: 18px; display: block; margin-bottom: 4px;}
.alert-success {background: var(--success-light); color: var(--success);}
.alert-warning {background: var(--warning-light); color: var(--warning);}
.alert-danger {background: var(--danger-light); color: var(--danger);}

.success-banner {
    display: flex; align-items: center; gap: 8px;
    padding: 10px 14px;
    background: var(--success-light);
    color: var(--success);
    font-size: 12px; font-weight: 600;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    margin-bottom: 16px;
}

.info-notice {
    display: flex; align-items: center; gap: 10px;
    padding: 12px 14px;
    background: var(--grey-warm);
    border-left: 3px solid var(--grey-text);
    font-size: 13px;
    color: var(--grey-text);
    margin-top: 12px;
}

.stSuccess, .stInfo {background: var(--success-light) !important; color: var(--success) !important; border-radius: 0 !important; border-left: 3px solid var(--success) !important;}
.stWarning {background: var(--warning-light) !important; color: var(--warning) !important; border-radius: 0 !important; border-left: 3px solid var(--warning) !important;}
.stError {background: var(--danger-light) !important; color: var(--danger) !important; border-radius: 0 !important; border-left: 3px solid var(--danger) !important;}
.streamlit-expanderHeader {background: var(--grey-warm) !important; border-radius: 0 !important;}
hr {border-color: var(--grey-border) !important;}

@media (max-width: 640px) {.alert-grid {grid-template-columns: repeat(2, 1fr);}}
</style>

<div class="hyab-header">
    <span class="brand-mark">‹</span>
    <span class="brand-text">HYAB</span>
    <span class="brand-divider">/</span>
    <span class="brand-page">Data Cleaner v2.1</span>
</div>
""", unsafe_allow_html=True)

def _page_title(text):
    st.markdown(f'<h1 class="page-title">{text}</h1>', unsafe_allow_html=True)

def _section_label(text):
    st.markdown(f'<p class="section-label">{text}</p>', unsafe_allow_html=True)

def _success_banner():
    st.markdown('<div class="success-banner">✓ Done</div>', unsafe_allow_html=True)

def _info_notice(text):
    st.markdown(f'<div class="info-notice">→ {text}</div>', unsafe_allow_html=True)

def _alert_grid(alerts):
    html = '<div class="alert-grid">'
    for a in alerts:
        html += f'<div class="alert-box alert-{a["status"]}"><span class="icon">{a["icon"]}</span>{a["text"]}</div>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

# --- Styles ---
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HIGHLIGHT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
ALERT_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
SUCCESS_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
GREEN_FONT = Font(color='006600', bold=True)
RED_FONT = Font(bold=True, color='CC0000')
BOLD_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
GREY_FONT = Font(color='808080', italic=True)

DEFAULT_FX_RATES = {'SEK': 1.0, 'EUR': 11.20, 'USD': 10.50, 'GBP': 13.30}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def find_sheet(wb, possible_names):
    """Find a sheet by trying multiple possible names (case-insensitive)."""
    sheet_names_lower = {name.lower(): name for name in wb.sheetnames}
    
    for name in possible_names:
        if name.lower() in sheet_names_lower:
            return wb[sheet_names_lower[name.lower()]]
    
    # Fallback: return first sheet if only one exists
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]
    
    return None


def clean_amount_with_currency(raw_value):
    """Parse amounts like '3 800.00 SEK' or '5 211.00 EUR'"""
    if raw_value is None:
        return None, None
    
    raw = str(raw_value).strip()
    match = re.match(r'([\d\s\xa0\.,]+)\s*(SEK|EUR|USD|GBP)?', raw, re.IGNORECASE)
    if not match:
        return None, None
    
    amount_str = match.group(1).replace('\xa0', '').replace(' ', '')
    currency = match.group(2).upper() if match.group(2) else 'SEK'
    
    if re.search(r',\d{2}$', amount_str):
        amount_str = amount_str.replace('.', '').replace(',', '.')
    else:
        amount_str = amount_str.replace(',', '')
    
    try:
        amount = float(amount_str)
    except ValueError:
        return None, None
    
    return amount, currency


def clean_number(raw_value):
    """Clean numbers with non-breaking spaces."""
    if raw_value is None:
        return None
    
    raw = str(raw_value).strip()
    if raw.endswith('%'):
        return raw
    if raw in ['', 'n/a', 'None', '-']:
        return None
    
    cleaned = raw.replace('\xa0', '').replace(' ', '')
    if re.search(r',\d{2}$', cleaned):
        cleaned = cleaned.replace('.', '').replace(',', '.')
    else:
        cleaned = cleaned.replace(',', '')
    
    try:
        return float(cleaned)
    except ValueError:
        return raw


def style_header(ws, headers, row=1):
    """Apply header styling."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def to_excel_bytes(workbook):
    """Convert workbook to bytes for download."""
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def format_sek(amount):
    """Format number as SEK string."""
    if abs(amount) >= 1_000_000:
        return f"{amount/1_000_000:.1f}M"
    elif abs(amount) >= 1_000:
        return f"{amount/1_000:.0f}k"
    else:
        return f"{amount:.0f}"


# =============================================================================
# ORDER BOOK PARSING (Updated for flexible sheet names)
# =============================================================================

def parse_order_book(ws_raw, fx_rates):
    """Parse raw order book into list of order dicts."""
    orders = []
    
    for row in range(2, ws_raw.max_row + 1):
        belopp, valuta = clean_amount_with_currency(ws_raw.cell(row, 6).value)
        if belopp is None:
            continue
        
        orderdatum = ws_raw.cell(row, 2).value
        if isinstance(orderdatum, datetime):
            order_date = orderdatum
        else:
            order_date = None
        
        rate = fx_rates.get(valuta, 1.0)
        belopp_sek = round(belopp * rate, 2)
        
        orders.append({
            'ordernr': ws_raw.cell(row, 1).value,
            'orderdatum': order_date,
            'kundnamn': ws_raw.cell(row, 3).value,
            'status': ws_raw.cell(row, 4).value,
            'fakt_status': ws_raw.cell(row, 5).value or '',
            'belopp': belopp,
            'valuta': valuta,
            'belopp_sek': belopp_sek,
        })
    
    return orders


# =============================================================================
# SALES PARSING (Updated for Victor's feedback)
# =============================================================================

def parse_sales_articles(ws_raw):
    """
    Parse Article sheet from raw sales data.
    Uses correct columns per Victor's feedback:
    - Col 1: Artikelnr
    - Col 2: Artikelnamn
    - Col 3: Summa utan moms
    - Col 6: Antal ut (correct quantity column)
    - Col 7: TB (keep but flag as unreliable)
    - Col 8: TG (keep but flag as unreliable)
    """
    articles = []
    
    for row in range(2, ws_raw.max_row + 1):
        artikelnr = ws_raw.cell(row, 1).value
        if artikelnr is None:
            continue
        
        summa = clean_number(ws_raw.cell(row, 3).value)
        if summa is None or summa == 0:
            continue
        
        articles.append({
            'artikelnr': str(artikelnr).strip(),
            'artikelnamn': ws_raw.cell(row, 2).value,
            'summa': summa,
            'antal_ut': clean_number(ws_raw.cell(row, 6).value) or 0,  # Correct column
            'tb': clean_number(ws_raw.cell(row, 7).value),
            'tg': ws_raw.cell(row, 8).value,  # Keep as string (percentage)
        })
    
    return articles


def parse_sales_customers(ws_raw):
    """
    Parse Company sheet from raw sales data.
    - Col 1: Kundnr
    - Col 2: Kund
    - Col 3: Kundtyp (ignore per Victor)
    - Col 4: Summa utan moms
    """
    customers = []
    
    for row in range(2, ws_raw.max_row + 1):
        kund = ws_raw.cell(row, 2).value
        if kund is None:
            continue
        
        summa = clean_number(ws_raw.cell(row, 4).value)
        if summa is None or summa == 0:
            continue
        
        customers.append({
            'kundnr': ws_raw.cell(row, 1).value,
            'kund': str(kund).strip(),
            'summa': summa,
        })
    
    return customers


def create_sales_output_v2(articles, customers, master_articles=None):
    """Create sales output with Top 20 analysis (v2)."""
    wb = openpyxl.Workbook()
    
    total_articles = sum(a['summa'] for a in articles)
    total_customers = sum(c['summa'] for c in customers)
    
    # Sort for Top 20
    articles_sorted = sorted(articles, key=lambda x: x['summa'], reverse=True)
    customers_sorted = sorted(customers, key=lambda x: x['summa'], reverse=True)
    
    top20_articles = articles_sorted[:20]
    top20_customers = customers_sorted[:20]
    
    top20_art_sum = sum(a['summa'] for a in top20_articles)
    top20_cust_sum = sum(c['summa'] for c in top20_customers)
    
    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    
    ws_summary['A1'] = 'HYAB MONTHLY SALES'
    ws_summary['A1'].font = TITLE_FONT
    ws_summary['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    
    row = 4
    ws_summary.cell(row, 1, 'OVERVIEW').font = SUBTITLE_FONT
    row += 1
    ws_summary.cell(row, 1, 'Total sales (articles):')
    ws_summary.cell(row, 2, f'{total_articles:,.0f} SEK')
    row += 1
    ws_summary.cell(row, 1, 'Total sales (customers):')
    ws_summary.cell(row, 2, f'{total_customers:,.0f} SEK')
    row += 1
    ws_summary.cell(row, 1, 'Article count:')
    ws_summary.cell(row, 2, len(articles))
    row += 1
    ws_summary.cell(row, 1, 'Customer count:')
    ws_summary.cell(row, 2, len(customers))
    row += 2
    
    ws_summary.cell(row, 1, 'CONCENTRATION').font = SUBTITLE_FONT
    row += 1
    ws_summary.cell(row, 1, 'Top 20 articles:')
    ws_summary.cell(row, 2, f'{top20_art_sum:,.0f} SEK ({top20_art_sum/total_articles*100:.1f}%)')
    row += 1
    ws_summary.cell(row, 1, 'Top 20 customers:')
    ws_summary.cell(row, 2, f'{top20_cust_sum:,.0f} SEK ({top20_cust_sum/total_customers*100:.1f}%)')
    row += 2
    
    # New articles to add
    new_articles = []
    if master_articles is not None:
        for a in articles:
            if a['artikelnr'].lower() not in master_articles:
                new_articles.append(a)
        
        if new_articles:
            ws_summary.cell(row, 1, f'🆕 {len(new_articles)} new articles to add to master')
            ws_summary.cell(row, 1).font = RED_FONT
            row += 1
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30
    
    # --- Sheet 2: Top 20 Articles ---
    ws_top_art = wb.create_sheet('Top 20 Articles')
    ws_top_art['A1'] = f'TOP 20 ARTICLES ({top20_art_sum/total_articles*100:.0f}% of total)'
    ws_top_art['A1'].font = TITLE_FONT
    
    row = 3
    style_header(ws_top_art, ['#', 'Article No', 'Article Name', 'Amount SEK', '% of Total', 'Qty'], row)
    row += 1
    
    for i, a in enumerate(top20_articles, 1):
        pct = (a['summa'] / total_articles * 100) if total_articles > 0 else 0
        ws_top_art.cell(row, 1, i)
        ws_top_art.cell(row, 2, a['artikelnr'])
        ws_top_art.cell(row, 3, a['artikelnamn'])
        ws_top_art.cell(row, 4, a['summa'])
        ws_top_art.cell(row, 5, f'{pct:.1f}%')
        ws_top_art.cell(row, 6, a['antal_ut'])
        row += 1
    
    row += 1
    ws_top_art.cell(row, 3, 'TOP 20 TOTAL')
    ws_top_art.cell(row, 3).font = BOLD_FONT
    ws_top_art.cell(row, 4, top20_art_sum)
    ws_top_art.cell(row, 4).font = BOLD_FONT
    
    ws_top_art.column_dimensions['A'].width = 5
    ws_top_art.column_dimensions['B'].width = 18
    ws_top_art.column_dimensions['C'].width = 40
    ws_top_art.column_dimensions['D'].width = 15
    ws_top_art.column_dimensions['E'].width = 10
    ws_top_art.column_dimensions['F'].width = 8
    
    # --- Sheet 3: Top 20 Customers ---
    ws_top_cust = wb.create_sheet('Top 20 Customers')
    ws_top_cust['A1'] = f'TOP 20 CUSTOMERS ({top20_cust_sum/total_customers*100:.0f}% of total)'
    ws_top_cust['A1'].font = TITLE_FONT
    
    row = 3
    style_header(ws_top_cust, ['#', 'Customer No', 'Customer Name', 'Amount SEK', '% of Total'], row)
    row += 1
    
    for i, c in enumerate(top20_customers, 1):
        pct = (c['summa'] / total_customers * 100) if total_customers > 0 else 0
        ws_top_cust.cell(row, 1, i)
        ws_top_cust.cell(row, 2, c['kundnr'])
        ws_top_cust.cell(row, 3, c['kund'])
        ws_top_cust.cell(row, 4, c['summa'])
        ws_top_cust.cell(row, 5, f'{pct:.1f}%')
        row += 1
    
    row += 1
    ws_top_cust.cell(row, 3, 'TOP 20 TOTAL')
    ws_top_cust.cell(row, 3).font = BOLD_FONT
    ws_top_cust.cell(row, 4, top20_cust_sum)
    ws_top_cust.cell(row, 4).font = BOLD_FONT
    
    ws_top_cust.column_dimensions['A'].width = 5
    ws_top_cust.column_dimensions['B'].width = 12
    ws_top_cust.column_dimensions['C'].width = 40
    ws_top_cust.column_dimensions['D'].width = 15
    ws_top_cust.column_dimensions['E'].width = 10
    
    # --- Sheet 4: Articles Raw Data ---
    ws_art = wb.create_sheet('Articles')
    headers = ['Article No', 'Article Name', 'Amount excl. VAT', 'Qty (ut)', 'TB*', 'TG*']
    style_header(ws_art, headers)
    
    row = 2
    for a in articles_sorted:
        ws_art.cell(row, 1, a['artikelnr'])
        ws_art.cell(row, 2, a['artikelnamn'])
        ws_art.cell(row, 3, a['summa'])
        ws_art.cell(row, 4, a['antal_ut'])
        ws_art.cell(row, 5, a['tb'])
        ws_art.cell(row, 6, a['tg'])
        row += 1
    
    row += 1
    ws_art.cell(row, 2, 'TOTAL')
    ws_art.cell(row, 2).font = BOLD_FONT
    ws_art.cell(row, 3, f'=SUM(C2:C{row-2})')
    
    row += 2
    ws_art.cell(row, 1, '* TB/TG values may be unreliable in current system')
    ws_art.cell(row, 1).font = GREY_FONT
    
    ws_art.column_dimensions['A'].width = 18
    ws_art.column_dimensions['B'].width = 40
    ws_art.column_dimensions['C'].width = 18
    ws_art.column_dimensions['D'].width = 10
    ws_art.column_dimensions['E'].width = 12
    ws_art.column_dimensions['F'].width = 10
    
    # --- Sheet 5: Customers Raw Data ---
    ws_cust = wb.create_sheet('Customers')
    headers = ['Customer No', 'Customer Name', 'Amount excl. VAT']
    style_header(ws_cust, headers)
    
    row = 2
    for c in customers_sorted:
        ws_cust.cell(row, 1, c['kundnr'])
        ws_cust.cell(row, 2, c['kund'])
        ws_cust.cell(row, 3, c['summa'])
        row += 1
    
    row += 1
    ws_cust.cell(row, 2, 'TOTAL')
    ws_cust.cell(row, 2).font = BOLD_FONT
    ws_cust.cell(row, 3, f'=SUM(C2:C{row-2})')
    
    ws_cust.column_dimensions['A'].width = 12
    ws_cust.column_dimensions['B'].width = 40
    ws_cust.column_dimensions['C'].width = 18
    
    # --- Sheet 6: New Articles (if master provided) ---
    if master_articles is not None:
        ws_new = wb.create_sheet('New Articles')
        style_header(ws_new, ['Article No', 'Article Name', 'Amount', 'Action'])
        
        if new_articles:
            row = 2
            for a in sorted(new_articles, key=lambda x: x['summa'], reverse=True):
                ws_new.cell(row, 1, a['artikelnr'])
                ws_new.cell(row, 2, a['artikelnamn'])
                ws_new.cell(row, 3, a['summa'])
                ws_new.cell(row, 4, 'ADD')
                ws_new.cell(row, 4).fill = HIGHLIGHT_FILL
                row += 1
        else:
            ws_new.cell(2, 1, '✓ No new articles')
            ws_new.cell(2, 1).font = GREEN_FONT
        
        ws_new.column_dimensions['A'].width = 18
        ws_new.column_dimensions['B'].width = 40
        ws_new.column_dimensions['C'].width = 12
        ws_new.column_dimensions['D'].width = 12
    
    stats = {
        'articles': len(articles),
        'customers': len(customers),
        'total': total_articles,
        'top20_art_pct': top20_art_sum / total_articles * 100 if total_articles > 0 else 0,
        'top20_cust_pct': top20_cust_sum / total_customers * 100 if total_customers > 0 else 0,
        'new_articles': new_articles,
        'top20_articles': top20_articles,
        'top20_customers': top20_customers,
    }
    
    return wb, stats


def load_master_articles(wb_master):
    """Load article list from master file."""
    articles = set()
    
    if 'Försäljning per artikel' in wb_master.sheetnames:
        ws = wb_master['Försäljning per artikel']
        for row in range(2, ws.max_row + 1):
            art = ws.cell(row, 1).value
            if art:
                articles.add(str(art).strip().lower())
    
    return articles


# =============================================================================
# ORDER BOOK ANALYTICS (Same as before)
# =============================================================================

def analyze_order_aging(orders, today=None):
    """Find orders that are getting stale (>3 months old)."""
    if today is None:
        today = datetime.now()
    
    aging_alerts = []
    for order in orders:
        if order['orderdatum'] is None:
            continue
        
        days_old = (today - order['orderdatum']).days
        if days_old > 90:
            aging_alerts.append({
                'ordernr': order['ordernr'],
                'kundnamn': order['kundnamn'],
                'orderdatum': order['orderdatum'],
                'days_old': days_old,
                'months_old': days_old // 30,
                'belopp_sek': order['belopp_sek'],
            })
    
    return sorted(aging_alerts, key=lambda x: x['days_old'], reverse=True)


def analyze_customer_concentration(orders):
    """Calculate customer concentration risk."""
    customer_totals = defaultdict(float)
    for order in orders:
        customer_totals[order['kundnamn']] += order['belopp_sek']
    
    total = sum(customer_totals.values())
    if total == 0:
        return [], 0, 0
    
    sorted_customers = sorted(customer_totals.items(), key=lambda x: x[1], reverse=True)
    top_5 = [(name, amt, amt/total*100) for name, amt in sorted_customers[:5]]
    top_3_pct = sum(amt for _, amt in sorted_customers[:3]) / total * 100
    top_1_pct = sorted_customers[0][1] / total * 100 if sorted_customers else 0
    
    return top_5, top_3_pct, top_1_pct


def analyze_week_over_week(current_orders, previous_orders):
    """Compare current week to previous week."""
    current_ids = {o['ordernr'] for o in current_orders}
    previous_ids = {o['ordernr'] for o in previous_orders}
    
    new_ids = current_ids - previous_ids
    new_orders = [o for o in current_orders if o['ordernr'] in new_ids]
    
    closed_ids = previous_ids - current_ids
    closed_orders = [o for o in previous_orders if o['ordernr'] in closed_ids]
    
    unchanged_ids = current_ids & previous_ids
    unchanged_orders = [o for o in current_orders if o['ordernr'] in unchanged_ids]
    
    new_total = sum(o['belopp_sek'] for o in new_orders)
    closed_total = sum(o['belopp_sek'] for o in closed_orders)
    unchanged_total = sum(o['belopp_sek'] for o in unchanged_orders)
    
    return {
        'new': {'count': len(new_orders), 'total': new_total, 'orders': new_orders},
        'closed': {'count': len(closed_orders), 'total': closed_total, 'orders': closed_orders},
        'unchanged': {'count': len(unchanged_orders), 'total': unchanged_total, 'orders': unchanged_orders},
        'net_count': len(new_orders) - len(closed_orders),
        'net_amount': new_total - closed_total,
    }


def analyze_large_orders(orders, threshold=100000):
    """Find orders above threshold."""
    return sorted(
        [o for o in orders if o['belopp_sek'] > threshold],
        key=lambda x: x['belopp_sek'],
        reverse=True
    )


# =============================================================================
# ORDER BOOK OUTPUT (Updated with delivery date column)
# =============================================================================

def create_order_book_output(orders, fx_rates, wow_data=None):
    """Create order book output with analytics."""
    wb = openpyxl.Workbook()
    
    total_sek = sum(o['belopp_sek'] for o in orders)
    aging = analyze_order_aging(orders)
    large = analyze_large_orders(orders)
    top_5, top_3_pct, _ = analyze_customer_concentration(orders)
    
    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    
    ws_summary['A1'] = 'HYAB ORDER BOOK'
    ws_summary['A1'].font = TITLE_FONT
    ws_summary['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    
    row = 4
    ws_summary.cell(row, 1, 'OVERVIEW').font = SUBTITLE_FONT
    row += 1
    ws_summary.cell(row, 1, 'Open orders:')
    ws_summary.cell(row, 2, len(orders))
    row += 1
    ws_summary.cell(row, 1, 'Total value:')
    ws_summary.cell(row, 2, f'{total_sek:,.0f} SEK')
    row += 2
    
    if wow_data:
        ws_summary.cell(row, 1, 'WEEK-OVER-WEEK').font = SUBTITLE_FONT
        row += 1
        ws_summary.cell(row, 1, 'New orders:')
        ws_summary.cell(row, 2, f"+{wow_data['new']['count']} ({format_sek(wow_data['new']['total'])} SEK)")
        ws_summary.cell(row, 2).font = GREEN_FONT
        row += 1
        ws_summary.cell(row, 1, 'Closed orders:')
        ws_summary.cell(row, 2, f"-{wow_data['closed']['count']} ({format_sek(wow_data['closed']['total'])} SEK)")
        row += 1
        ws_summary.cell(row, 1, 'Net change:')
        net_str = f"{wow_data['net_count']:+d} orders, {format_sek(wow_data['net_amount'])} SEK"
        ws_summary.cell(row, 2, net_str)
        ws_summary.cell(row, 2).font = GREEN_FONT if wow_data['net_amount'] >= 0 else RED_FONT
        row += 2
    
    ws_summary.cell(row, 1, 'ALERTS').font = SUBTITLE_FONT
    row += 1
    if aging:
        ws_summary.cell(row, 1, f'⏰ {len(aging)} orders older than 3 months')
        ws_summary.cell(row, 1).font = RED_FONT
        row += 1
    if large:
        ws_summary.cell(row, 1, f'💰 {len(large)} orders over 100k SEK')
        row += 1
    if top_3_pct > 50:
        ws_summary.cell(row, 1, f'📊 Top 3 customers = {top_3_pct:.0f}% of value')
        row += 1
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 30
    
    # --- Sheet 2: Week-over-Week ---
    if wow_data:
        ws_wow = wb.create_sheet('Week-over-Week')
        row = 1
        ws_wow.cell(row, 1, 'NEW ORDERS').font = SUBTITLE_FONT
        row += 1
        
        if wow_data['new']['orders']:
            style_header(ws_wow, ['Order No', 'Customer', 'Amount SEK'], row)
            row += 1
            for o in wow_data['new']['orders'][:20]:
                ws_wow.cell(row, 1, o['ordernr'])
                ws_wow.cell(row, 2, o['kundnamn'])
                ws_wow.cell(row, 3, o['belopp_sek'])
                ws_wow.cell(row, 1).fill = SUCCESS_FILL
                row += 1
        else:
            ws_wow.cell(row, 1, 'No new orders')
            row += 1
        
        row += 1
        ws_wow.cell(row, 1, 'CLOSED ORDERS').font = SUBTITLE_FONT
        row += 1
        
        if wow_data['closed']['orders']:
            style_header(ws_wow, ['Order No', 'Customer', 'Amount SEK'], row)
            row += 1
            for o in wow_data['closed']['orders'][:20]:
                ws_wow.cell(row, 1, o['ordernr'])
                ws_wow.cell(row, 2, o['kundnamn'])
                ws_wow.cell(row, 3, o['belopp_sek'])
                row += 1
        
        ws_wow.column_dimensions['A'].width = 12
        ws_wow.column_dimensions['B'].width = 35
        ws_wow.column_dimensions['C'].width = 15
    
    # --- Sheet 3: Warnings ---
    ws_alerts = wb.create_sheet('Warnings')
    ws_alerts['A1'] = 'WARNINGS'
    ws_alerts['A1'].font = TITLE_FONT
    row = 3
    
    if aging:
        ws_alerts.cell(row, 1, 'AGING ORDERS (>3 months)').font = SUBTITLE_FONT
        row += 1
        style_header(ws_alerts, ['Order No', 'Customer', 'Order Date', 'Months', 'Amount SEK'], row)
        row += 1
        for a in aging[:15]:
            ws_alerts.cell(row, 1, a['ordernr'])
            ws_alerts.cell(row, 2, a['kundnamn'])
            ws_alerts.cell(row, 3, a['orderdatum'])
            ws_alerts.cell(row, 4, f"{a['months_old']} mo")
            ws_alerts.cell(row, 5, a['belopp_sek'])
            ws_alerts.cell(row, 1).fill = ALERT_FILL
            row += 1
        row += 1
    
    if large:
        ws_alerts.cell(row, 1, 'LARGE ORDERS (>100k SEK)').font = SUBTITLE_FONT
        row += 1
        style_header(ws_alerts, ['Order No', 'Customer', 'Amount SEK', 'Currency'], row)
        row += 1
        for o in large[:10]:
            ws_alerts.cell(row, 1, o['ordernr'])
            ws_alerts.cell(row, 2, o['kundnamn'])
            ws_alerts.cell(row, 3, o['belopp_sek'])
            ws_alerts.cell(row, 4, o['valuta'])
            ws_alerts.cell(row, 1).fill = HIGHLIGHT_FILL
            row += 1
    
    ws_alerts.column_dimensions['A'].width = 12
    ws_alerts.column_dimensions['B'].width = 35
    ws_alerts.column_dimensions['C'].width = 15
    ws_alerts.column_dimensions['D'].width = 12
    ws_alerts.column_dimensions['E'].width = 15
    
    # --- Sheet 4: Order Book (with delivery date placeholder) ---
    ws_data = wb.create_sheet('Order Book')
    
    # Updated headers with Delivery Date column
    headers = ['Order No', 'Order Date', 'Customer', 'Status', 'Inv.status', 
               'Amount', 'Currency', 'Amount SEK', 'Delivery Date', 'Month', 
               'Count art.', 'Risk', 'Risk amount', 'Comment']
    style_header(ws_data, headers)
    
    # Highlight manual-entry columns
    for col in [9, 10, 11, 12, 13, 14]:
        ws_data.cell(1, col).fill = HIGHLIGHT_FILL
    
    row = 2
    for o in orders:
        ws_data.cell(row, 1, o['ordernr'])
        ws_data.cell(row, 2, o['orderdatum'])
        ws_data.cell(row, 3, o['kundnamn'])
        ws_data.cell(row, 4, o['status'])
        ws_data.cell(row, 5, o['fakt_status'])
        ws_data.cell(row, 6, o['belopp'])
        ws_data.cell(row, 7, o['valuta'])
        ws_data.cell(row, 8, o['belopp_sek'])
        # Col 9 (Delivery Date) - leave empty for manual entry
        ws_data.cell(row, 10, f'=IF(I{row}="","",DATE(YEAR(I{row}),MONTH(I{row}),1))')
        
        if 'fakturerad' in str(o['fakt_status']).lower() or 'partial' in str(o['fakt_status']).lower():
            for col in range(1, 15):
                ws_data.cell(row, col).fill = HIGHLIGHT_FILL
        
        row += 1
    
    row += 1
    ws_data.cell(row, 1, 'TOTAL').font = BOLD_FONT
    ws_data.cell(row, 8, f'=SUM(H2:H{row-2})')
    ws_data.cell(row, 8).font = BOLD_FONT
    
    row += 2
    ws_data.cell(row, 1, 'Exchange Rates:').font = BOLD_FONT
    for curr, rate in fx_rates.items():
        if curr != 'SEK':
            row += 1
            ws_data.cell(row, 1, f'{curr}:')
            ws_data.cell(row, 2, rate)
    
    row += 2
    ws_data.cell(row, 1, 'Yellow columns = manual entry required').font = GREY_FONT
    
    widths = {'A': 10, 'B': 12, 'C': 30, 'D': 10, 'E': 12, 'F': 10, 'G': 8, 
              'H': 12, 'I': 14, 'J': 12, 'K': 10, 'L': 8, 'M': 12, 'N': 20}
    for col, width in widths.items():
        ws_data.column_dimensions[col].width = width
    
    # --- Sheet 5: By Delivery Month (pivot placeholder) ---
    ws_monthly = wb.create_sheet('By Delivery Month')
    ws_monthly['A1'] = 'ORDER BOOK BY DELIVERY MONTH'
    ws_monthly['A1'].font = TITLE_FONT
    ws_monthly['A2'] = '(Fill in delivery dates in "Order Book" sheet to populate)'
    ws_monthly['A2'].font = GREY_FONT
    
    ws_monthly['A4'] = 'Month'
    ws_monthly['B4'] = 'Order Count'
    ws_monthly['C4'] = 'Total SEK'
    style_header(ws_monthly, ['Month', 'Order Count', 'Total SEK'], 4)
    
    ws_monthly.column_dimensions['A'].width = 15
    ws_monthly.column_dimensions['B'].width = 15
    ws_monthly.column_dimensions['C'].width = 15
    
    return wb


# =============================================================================
# SALES INTELLIGENCE (Keep existing functionality)
# =============================================================================

def parse_sales_master(wb):
    """Parse Sales_work_file.xlsx master file for Intelligence mode."""
    if 'Försäljning per kund' not in wb.sheetnames:
        return None, "Could not find 'Försäljning per kund' sheet"
    
    ws = wb['Försäljning per kund']
    
    ltm_now_col = None
    ltm_prev_col = None
    
    for col in range(80, min(ws.max_column + 1, 100)):
        header = ws.cell(1, col).value
        if header:
            header_str = str(header).lower()
            if 'ltm' in header_str and '25' in header_str:
                ltm_now_col = col
            elif 'ltm' in header_str and '24' in header_str and ltm_prev_col is None:
                ltm_prev_col = col
    
    if ltm_now_col is None:
        ltm_now_col = 89
    if ltm_prev_col is None:
        ltm_prev_col = 78
    
    customers = []
    for row in range(2, ws.max_row + 1):
        kund = ws.cell(row, 2).value
        if not kund or kund == 'Summa':
            continue
        
        ltm_now = ws.cell(row, ltm_now_col).value or 0
        ltm_prev = ws.cell(row, ltm_prev_col).value or 0
        
        if not isinstance(ltm_now, (int, float)):
            ltm_now = 0
        if not isinstance(ltm_prev, (int, float)):
            ltm_prev = 0
        
        customers.append({
            'kund': kund,
            'ltm_now': ltm_now,
            'ltm_prev': ltm_prev,
            'change': ltm_now - ltm_prev,
        })
    
    return customers, None


def analyze_sales_intelligence(customers):
    """Analyze customer data for intelligence report."""
    total_now = sum(c['ltm_now'] for c in customers)
    total_prev = sum(c['ltm_prev'] for c in customers)
    total_change = total_now - total_prev
    pct_change = (total_change / total_prev * 100) if total_prev > 0 else 0
    
    active = [c for c in customers if c['ltm_now'] > 0]
    churned = [c for c in customers if c['ltm_prev'] > 50000 and c['ltm_now'] == 0]
    declining = [c for c in customers if c['ltm_prev'] > c['ltm_now'] > 0 and c['change'] < -20000]
    growing = [c for c in customers if c['change'] > 20000]
    new_customers = [c for c in customers if c['ltm_prev'] == 0 and c['ltm_now'] > 10000]
    
    churn_revenue = sum(c['ltm_prev'] for c in churned)
    decline_revenue = sum(abs(c['change']) for c in declining)
    growth_revenue = sum(c['change'] for c in growing)
    new_revenue = sum(c['ltm_now'] for c in new_customers)
    
    sorted_cust = sorted(active, key=lambda x: x['ltm_now'], reverse=True)
    top10_rev = sum(c['ltm_now'] for c in sorted_cust[:10])
    top20_rev = sum(c['ltm_now'] for c in sorted_cust[:20])
    
    return {
        'total_now': total_now,
        'total_prev': total_prev,
        'total_change': total_change,
        'pct_change': pct_change,
        'active': active,
        'churned': sorted(churned, key=lambda x: x['ltm_prev'], reverse=True),
        'declining': sorted(declining, key=lambda x: x['change']),
        'growing': sorted(growing, key=lambda x: x['change'], reverse=True),
        'new_customers': sorted(new_customers, key=lambda x: x['ltm_now'], reverse=True),
        'churn_revenue': churn_revenue,
        'decline_revenue': decline_revenue,
        'growth_revenue': growth_revenue,
        'new_revenue': new_revenue,
        'top10_pct': (top10_rev / total_now * 100) if total_now > 0 else 0,
        'top20_pct': (top20_rev / total_now * 100) if total_now > 0 else 0,
        'top20': sorted_cust[:20],
    }


# =============================================================================
# MAIN APP
# =============================================================================

_page_title("Clean your data")

_section_label("Mode")
mode = st.radio("Mode", ["Order Book", "Sales", "Intelligence"], label_visibility="collapsed")

st.markdown("<div style='height: 16px'></div>", unsafe_allow_html=True)


# =============================================================================
# ORDER BOOK MODE
# =============================================================================

if mode == "Order Book":
    _section_label("Exchange rates")
    col1, col2, col3 = st.columns(3)
    with col1:
        eur_rate = st.number_input("EUR", value=DEFAULT_FX_RATES['EUR'], step=0.01, format="%.2f")
    with col2:
        usd_rate = st.number_input("USD", value=DEFAULT_FX_RATES['USD'], step=0.01, format="%.2f")
    with col3:
        gbp_rate = st.number_input("GBP", value=DEFAULT_FX_RATES['GBP'], step=0.01, format="%.2f")
    
    fx_rates = {'SEK': 1.0, 'EUR': eur_rate, 'USD': usd_rate, 'GBP': gbp_rate}
    
    st.markdown("<div style='height: 16px'></div>", unsafe_allow_html=True)
    
    _section_label("Files")
    col1, col2 = st.columns(2)
    with col1:
        current_file = st.file_uploader("This week", type=['xlsx'], key="ob_current")
    with col2:
        previous_file = st.file_uploader("Last week (optional)", type=['xlsx'], key="ob_previous")
    
    st.markdown("<div style='height: 8px'></div>", unsafe_allow_html=True)
    
    if st.button("Process", type="primary", disabled=not current_file, use_container_width=True):
        if current_file:
            try:
                wb_current = openpyxl.load_workbook(current_file)
                
                # Flexible sheet name detection
                ws = find_sheet(wb_current, ['Order book', 'Order_book', 'Orderbook', 'Sheet1', 'Orders'])
                
                if ws is None:
                    st.error(f"Could not find order book sheet. Available sheets: {wb_current.sheetnames}")
                else:
                    with st.spinner("Processing..."):
                        current_orders = parse_order_book(ws, fx_rates)
                        
                        wow_data = None
                        if previous_file:
                            try:
                                wb_prev = openpyxl.load_workbook(previous_file)
                                ws_prev = find_sheet(wb_prev, ['Order book', 'Order_book', 'Orderbook', 'Sheet1', 'Orders'])
                                if ws_prev:
                                    previous_orders = parse_order_book(ws_prev, fx_rates)
                                    wow_data = analyze_week_over_week(current_orders, previous_orders)
                            except Exception:
                                pass
                        
                        wb_out = create_order_book_output(current_orders, fx_rates, wow_data)
                    
                    st.session_state['ob_results'] = {
                        'orders': current_orders,
                        'wow_data': wow_data,
                        'workbook': to_excel_bytes(wb_out),
                    }
                    st.rerun()
            
            except Exception as e:
                st.error(f"Error: {str(e)}")
    
    if 'ob_results' in st.session_state:
        results = st.session_state['ob_results']
        orders = results['orders']
        wow_data = results['wow_data']
        
        _success_banner()
        
        total_sek = sum(o['belopp_sek'] for o in orders)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Orders", len(orders))
        with col2:
            st.metric("Total SEK", f"{total_sek:,.0f}")
        with col3:
            if wow_data:
                delta_str = f"{wow_data['net_count']:+d} orders"
                st.metric("vs last week", f"{format_sek(wow_data['net_amount'])}", delta_str)
            else:
                st.metric("vs last week", "—")
        
        st.markdown("<div style='height: 16px'></div>", unsafe_allow_html=True)
        _section_label("Warnings")
        
        aging = analyze_order_aging(orders)
        large = analyze_large_orders(orders)
        top_5, top_3_pct, _ = analyze_customer_concentration(orders)
        partial_inv = [o for o in orders if 'fakturerad' in str(o['fakt_status']).lower()]
        
        alerts = []
        alerts.append({'icon': '⏰' if aging else '✓', 'text': f'{len(aging)} older >3mo' if aging else 'No old orders', 'status': 'warning' if aging else 'success'})
        alerts.append({'icon': '💰' if large else '✓', 'text': f'{len(large)} >100k' if large else 'No large orders', 'status': 'warning' if large else 'success'})
        alerts.append({'icon': '📊' if top_3_pct > 50 else '✓', 'text': f'Top3 = {top_3_pct:.0f}%' if top_3_pct > 50 else 'Diversified', 'status': 'warning' if top_3_pct > 50 else 'success'})
        alerts.append({'icon': '📝' if partial_inv else '✓', 'text': f'{len(partial_inv)} partial inv' if partial_inv else 'No partial inv', 'status': 'warning' if partial_inv else 'success'})
        
        _alert_grid(alerts)
        
        st.download_button(
            label="Download report",
            data=results['workbook'],
            file_name=f"HYAB_Order_Book_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


# =============================================================================
# SALES MODE (Updated with Top 20 analysis)
# =============================================================================

elif mode == "Sales":
    _section_label("Files")
    col1, col2 = st.columns(2)
    with col1:
        raw_file = st.file_uploader("This month raw data", type=['xlsx'], key="sales_raw")
    with col2:
        master_file = st.file_uploader("Master file (optional)", type=['xlsx'], key="sales_master")
    
    st.markdown("<div style='height: 8px'></div>", unsafe_allow_html=True)
    
    if st.button("Process", type="primary", disabled=not raw_file, use_container_width=True):
        if raw_file:
            try:
                wb_raw = openpyxl.load_workbook(raw_file)
                
                # Find Article and Company sheets
                ws_article = find_sheet(wb_raw, ['Article', 'Articles', 'Artikel', 'Sales'])
                ws_company = find_sheet(wb_raw, ['Company', 'Customer', 'Customers', 'Kund', 'Kunder'])
                
                if ws_article is None:
                    st.error(f"Could not find Article sheet. Available sheets: {wb_raw.sheetnames}")
                else:
                    with st.spinner("Processing..."):
                        articles = parse_sales_articles(ws_article)
                        customers = parse_sales_customers(ws_company) if ws_company else []
                        
                        master_articles = None
                        if master_file:
                            wb_master = openpyxl.load_workbook(master_file, data_only=True)
                            master_articles = load_master_articles(wb_master)
                        
                        wb_out, stats = create_sales_output_v2(articles, customers, master_articles)
                    
                    st.session_state['sales_results'] = {
                        'stats': stats,
                        'workbook': to_excel_bytes(wb_out),
                        'articles': articles,
                        'customers': customers,
                    }
                    st.rerun()
            
            except Exception as e:
                st.error(f"Error: {str(e)}")
    
    if 'sales_results' in st.session_state:
        results = st.session_state['sales_results']
        stats = results['stats']
        
        _success_banner()
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Sales", f"{stats['total']:,.0f} SEK")
        with col2:
            st.metric("Articles", stats['articles'])
        with col3:
            st.metric("Customers", stats['customers'])
        
        st.markdown("<div style='height: 16px'></div>", unsafe_allow_html=True)
        
        # Top 20 charts
        _section_label("Concentration Analysis")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Top 20 Articles Pie
            top20_art = stats['top20_articles']
            rest_art = stats['total'] - sum(a['summa'] for a in top20_art)
            
            fig_art = go.Figure(data=[go.Pie(
                labels=['Top 20 Articles', 'Other Articles'],
                values=[sum(a['summa'] for a in top20_art), rest_art],
                hole=0.5,
                marker_colors=['#1F4E79', '#E2E0DC'],
                textinfo='percent',
                textposition='outside',
            )])
            fig_art.update_layout(
                title=dict(text=f"Articles: Top 20 = {stats['top20_art_pct']:.0f}%", font=dict(size=13)),
                height=250,
                margin=dict(l=20, r=20, t=40, b=20),
                showlegend=False,
            )
            st.plotly_chart(fig_art, use_container_width=True)
        
        with col2:
            # Top 20 Customers Pie
            top20_cust = stats['top20_customers']
            total_cust = sum(c['summa'] for c in results['customers']) if results['customers'] else stats['total']
            rest_cust = total_cust - sum(c['summa'] for c in top20_cust)
            
            fig_cust = go.Figure(data=[go.Pie(
                labels=['Top 20 Customers', 'Other Customers'],
                values=[sum(c['summa'] for c in top20_cust), rest_cust],
                hole=0.5,
                marker_colors=['#4A9BA8', '#E2E0DC'],
                textinfo='percent',
                textposition='outside',
            )])
            fig_cust.update_layout(
                title=dict(text=f"Customers: Top 20 = {stats['top20_cust_pct']:.0f}%", font=dict(size=13)),
                height=250,
                margin=dict(l=20, r=20, t=40, b=20),
                showlegend=False,
            )
            st.plotly_chart(fig_cust, use_container_width=True)
        
        # Alerts
        if stats['new_articles']:
            st.markdown(f'<div class="alert-box alert-danger" style="margin: 16px 0; text-align: center;"><span class="icon">🆕</span> {len(stats["new_articles"])} new articles to add</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-box alert-success" style="margin: 16px 0; text-align: center;"><span class="icon">✓</span> All articles matched</div>', unsafe_allow_html=True)
        
        # Top 5 preview
        with st.expander("Top 5 Articles"):
            for i, a in enumerate(stats['top20_articles'][:5], 1):
                pct = (a['summa'] / stats['total'] * 100)
                st.markdown(f"**{i}. {a['artikelnr']}** - {a['artikelnamn'][:40]} — {a['summa']:,.0f} SEK ({pct:.1f}%)")
        
        with st.expander("Top 5 Customers"):
            for i, c in enumerate(stats['top20_customers'][:5], 1):
                total_cust = sum(x['summa'] for x in results['customers']) if results['customers'] else stats['total']
                pct = (c['summa'] / total_cust * 100) if total_cust > 0 else 0
                st.markdown(f"**{i}. {c['kund']}** — {c['summa']:,.0f} SEK ({pct:.1f}%)")
        
        st.download_button(
            label="Download report",
            data=results['workbook'],
            file_name=f"HYAB_Sales_{datetime.now().strftime('%Y%m')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


# =============================================================================
# INTELLIGENCE MODE (Keep existing - abbreviated for space)
# =============================================================================

else:
    st.markdown("""
    <div style="background: #E8F4FD; border-left: 3px solid #1F4E79; padding: 12px 16px; margin-bottom: 16px;">
        <strong>Sales Intelligence Report</strong><br>
        <span style="font-size: 13px; color: #666;">
        Upload your Sales_work_file.xlsx to generate customer health analysis.
        </span>
    </div>
    """, unsafe_allow_html=True)
    
    _section_label("Master File")
    master_file = st.file_uploader("Sales_work_file.xlsx", type=['xlsx'], key="intel_master")
    
    if st.button("Generate Report", type="primary", disabled=not master_file, use_container_width=True):
        if master_file:
            try:
                with st.spinner("Analyzing..."):
                    wb = openpyxl.load_workbook(master_file, data_only=True)
                    customers, error = parse_sales_master(wb)
                    
                    if error:
                        st.error(error)
                    else:
                        analysis = analyze_sales_intelligence(customers)
                        st.session_state['intel_results'] = {'analysis': analysis}
                        st.rerun()
            except Exception as e:
                st.error(f"Error: {str(e)}")
    
    if 'intel_results' in st.session_state:
        analysis = st.session_state['intel_results']['analysis']
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("LTM Sales", f"{analysis['total_now']/1_000_000:.1f}M")
        with col2:
            st.metric("YoY Change", f"{analysis['pct_change']:+.1f}%")
        with col3:
            st.metric("Active", f"{len(analysis['active'])}")
        with col4:
            st.metric("Churned", f"{len(analysis['churned'])}")
        
        st.markdown("---")
        
        st.markdown(f"""
        **Summary:** {len(analysis['churned'])} customers churned ({format_sek(analysis['churn_revenue'])} lost), 
        {len(analysis['declining'])} declining, {len(analysis['growing'])} growing, 
        {len(analysis['new_customers'])} new.
        """)


# =============================================================================
# FOOTER
# =============================================================================

st.markdown("---")

with st.expander("What's new in v2.1"):
    st.markdown("""
**Based on Victor's feedback:**
- ✅ Fixed sheet name detection (accepts "Sheet1", etc.)
- ✅ Sales: Added Top 20 Articles & Customers analysis
- ✅ Sales: Uses correct quantity column (Antal ut)
- ✅ Sales: Processes both Article and Company sheets
- ✅ Sales: Flags TB/TG as potentially unreliable
- ✅ Order Book: Added delivery date placeholder column

**Coming next:**
- Order book sum by delivery month
- Month-end historical tracking
- More period filtering options
    """)
