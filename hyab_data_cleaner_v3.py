"""
HYAB Data Cleaner + Intelligence Dashboard v3.0
Generates downloadable HTML dashboard with all of Victor's requested charts

Features:
- Order Book: Cleaning, FX conversion, aging alerts
- Sales: Monthly cleanup with Top 20 analysis
- Intelligence: Generates beautiful HTML dashboard with:
  - LTM trend chart (rolling 12 months over time)
  - Monthly sales bar chart
  - YoY same-month comparison
  - Revenue bridge visualization
  - Customer cohorts (Churned, Declining, Growing, New)
  - Top 20 Customers AND Articles
  - Concentration analysis
"""

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from datetime import datetime
from io import BytesIO
from collections import defaultdict
import json

st.set_page_config(page_title="HYAB Data Cleaner", page_icon="📊", layout="centered", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Libre+Baskerville:ital@1&family=Source+Sans+3:wght@400;500;600&display=swap');

:root {
    --navy: #1B2A41;
    --navy-dark: #141F30;
    --grey-warm: #F7F6F4;
    --grey-border: #E2E0DC;
    --grey-text: #6B7280;
    --teal: #4A9BA8;
    --success: #2D6A4F;
    --success-light: #D8F3DC;
}

#MainMenu, footer, header, .stDeployButton {display: none !important;}
section[data-testid="stSidebar"] {display: none;}
.stApp {background: var(--grey-warm);}

/* Center the main content */
.main .block-container {
    max-width: 900px;
    padding-top: 1rem;
    padding-left: 2rem;
    padding-right: 2rem;
}

.hyab-header {
    background: var(--navy);
    padding: 14px 24px;
    margin: -1rem -2rem 1.5rem -2rem;
    display: flex;
    align-items: center;
    gap: 12px;
    font-family: 'Source Sans 3', sans-serif;
}
.hyab-header .brand-mark {color: var(--teal); font-size: 22px;}
.hyab-header .brand-text {color: white; font-size: 13px; font-weight: 600; letter-spacing: 0.12em;}
.hyab-header .brand-divider {color: rgba(255,255,255,0.25); margin: 0 4px;}
.hyab-header .brand-page {color: rgba(255,255,255,0.7); font-size: 13px;}

.page-title {
    font-family: 'Libre Baskerville', Georgia, serif;
    font-size: 26px;
    font-style: italic;
    font-weight: 400;
    color: var(--navy);
    margin: 0 0 20px 0;
}

.section-label {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--grey-text);
    margin: 0 0 10px 0;
}

div[data-testid="stRadio"] > div {
    flex-direction: row !important;
    gap: 0 !important;
    background: white;
    border: 1px solid var(--grey-border);
    padding: 3px;
    width: fit-content;
}
div[data-testid="stRadio"] > div > label {
    padding: 8px 18px !important;
    margin: 0 !important;
    border-radius: 0 !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    background: transparent !important;
}
div[data-testid="stRadio"] > div > label[data-checked="true"] {
    background: var(--navy) !important;
    color: white !important;
}
div[data-testid="stRadio"] > label {display: none !important;}

.stButton > button {
    background: var(--navy) !important;
    color: white !important;
    border: none !important;
    border-radius: 0 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    height: 48px !important;
    width: 100% !important;
}
.stButton > button:hover {background: var(--navy-dark) !important;}
.stButton > button:disabled {background: var(--grey-border) !important; color: var(--grey-text) !important;}

.stDownloadButton > button {
    background: var(--success) !important;
    color: white !important;
    border: none !important;
    border-radius: 0 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    height: 48px !important;
}

div[data-testid="stMetric"] {
    background: white;
    padding: 16px;
    border-left: 3px solid var(--navy);
    border-radius: 0;
}
div[data-testid="stMetric"] label {
    font-size: 11px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
    color: var(--grey-text) !important;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    font-size: 24px !important;
    font-weight: 600 !important;
    color: var(--navy) !important;
}

.success-banner {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 10px 14px;
    background: var(--success-light);
    color: var(--success);
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    margin-bottom: 16px;
}

hr {border-color: var(--grey-border) !important;}
</style>

<div class="hyab-header">
    <span class="brand-mark">‹</span>
    <span class="brand-text">HYAB</span>
    <span class="brand-divider">/</span>
    <span class="brand-page">Data Cleaner v3.0</span>
</div>
""", unsafe_allow_html=True)

def _page_title(t): st.markdown(f'<h1 class="page-title">{t}</h1>', unsafe_allow_html=True)
def _section_label(t): st.markdown(f'<p class="section-label">{t}</p>', unsafe_allow_html=True)
def _success_banner(): st.markdown('<div class="success-banner">✓ Done</div>', unsafe_allow_html=True)

DEFAULT_FX = {'SEK': 1.0, 'EUR': 11.20, 'USD': 10.50, 'GBP': 13.30}

def find_sheet(wb, names):
    lower = {n.lower(): n for n in wb.sheetnames}
    for name in names:
        if name.lower() in lower: return wb[lower[name.lower()]]
    return wb[wb.sheetnames[0]] if len(wb.sheetnames) == 1 else None

def clean_amount(raw):
    if raw is None: return None, None
    m = re.match(r'([\d\s\xa0\.,]+)\s*(SEK|EUR|USD|GBP)?', str(raw).strip(), re.IGNORECASE)
    if not m: return None, None
    s = m.group(1).replace('\xa0', '').replace(' ', '')
    cur = m.group(2).upper() if m.group(2) else 'SEK'
    if re.search(r',\d{2}$', s): s = s.replace('.', '').replace(',', '.')
    else: s = s.replace(',', '')
    try: return float(s), cur
    except: return None, None

def clean_num(raw):
    if raw is None: return None
    s = str(raw).strip()
    if s in ['', 'n/a', 'None', '-']: return None
    s = s.replace('\xa0', '').replace(' ', '')
    if re.search(r',\d{2}$', s): s = s.replace('.', '').replace(',', '.')
    else: s = s.replace(',', '')
    try: return float(s)
    except: return raw

def fmt_sek(n):
    if abs(n) >= 1e6: return f"{n/1e6:.1f}M"
    if abs(n) >= 1e3: return f"{n/1e3:.0f}k"
    return f"{n:.0f}"

def fmt_num(n): return f"{n:,.0f}".replace(",", " ")

def parse_master(wb):
    data = {'articles': [], 'customers': [], 'monthly_totals': {}, 'ltm_trend': {}}
    
    if 'Försäljning per artikel' in wb.sheetnames:
        ws = wb['Försäljning per artikel']
        months, ltms, fys = {}, {}, {}
        for c in range(3, min(ws.max_column + 1, 100)):
            h = ws.cell(1, c).value
            if h is None: continue
            hs = str(h)
            if isinstance(h, datetime): months[h.strftime('%Y-%m')] = c
            elif hs.startswith('FY'): fys[hs] = c
            elif hs.startswith('LTM'): ltms[hs] = c
            elif hs == 'YTD': fys['YTD'] = c
        
        for r in range(2, ws.max_row + 1):
            art = ws.cell(r, 1).value
            if not art or art == 'Summa': continue
            a = {'artikelnr': str(art), 'artikelnamn': ws.cell(r, 2).value or '', 'monthly': {}, 'fy': {}, 'ltm': {}}
            for k, c in months.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): a['monthly'][k] = v
            for k, c in fys.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): a['fy'][k] = v
            for k, c in ltms.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): a['ltm'][k] = v
            if any(a['monthly'].values()) or any(a['fy'].values()): data['articles'].append(a)
    
    if 'Försäljning per kund' in wb.sheetnames:
        ws = wb['Försäljning per kund']
        months, ltms, fys, churn_c = {}, {}, {}, None
        for c in range(3, min(ws.max_column + 1, 100)):
            h = ws.cell(1, c).value
            if h is None: continue
            hs = str(h)
            if isinstance(h, datetime): months[h.strftime('%Y-%m')] = c
            elif hs.startswith('FY'): fys[hs] = c
            elif hs.startswith('LTM'): ltms[hs] = c
            elif hs == 'YTD': fys['YTD'] = c
            elif 'bortfall' in hs.lower(): churn_c = c
        
        for r in range(2, ws.max_row + 1):
            kund = ws.cell(r, 2).value
            if not kund or kund == 'Summa': continue
            cust = {'kund': str(kund), 'monthly': {}, 'fy': {}, 'ltm': {}, 'churned': False}
            if churn_c:
                cv = ws.cell(r, churn_c).value
                if cv and str(cv).strip(): cust['churned'] = True
            for k, c in months.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): cust['monthly'][k] = v
            for k, c in fys.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): cust['fy'][k] = v
            for k, c in ltms.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): cust['ltm'][k] = v
            if any(cust['monthly'].values()) or any(cust['fy'].values()): data['customers'].append(cust)
    
    all_m = set()
    for a in data['articles']: all_m.update(a['monthly'].keys())
    for m in sorted(all_m): data['monthly_totals'][m] = sum(a['monthly'].get(m, 0) for a in data['articles'])
    
    all_l = set()
    for c in data['customers']: all_l.update(c['ltm'].keys())
    for l in sorted(all_l): data['ltm_trend'][l] = sum(c['ltm'].get(l, 0) for c in data['customers'])
    
    return data

def analyze_cohorts(data, curr, prev):
    churned, declining, growing, new = [], [], [], []
    for c in data['customers']:
        cur_v = c['ltm'].get(curr, 0)
        pre_v = c['ltm'].get(prev, 0)
        chg = cur_v - pre_v
        if c['churned'] or (pre_v > 0 and cur_v == 0):
            churned.append({'kund': c['kund'], 'previous': pre_v, 'current': cur_v, 'change': chg})
        elif pre_v == 0 and cur_v > 0:
            new.append({'kund': c['kund'], 'current': cur_v})
        elif chg < 0:
            declining.append({'kund': c['kund'], 'previous': pre_v, 'current': cur_v, 'change': chg, 'pct': (chg/pre_v*100) if pre_v > 0 else 0})
        elif chg > 0:
            growing.append({'kund': c['kund'], 'previous': pre_v, 'current': cur_v, 'change': chg, 'pct': (chg/pre_v*100) if pre_v > 0 else 0})
    
    churned.sort(key=lambda x: x['previous'], reverse=True)
    declining.sort(key=lambda x: x['change'])
    growing.sort(key=lambda x: x['change'], reverse=True)
    new.sort(key=lambda x: x['current'], reverse=True)
    return {'churned': churned, 'declining': declining, 'growing': growing, 'new': new}

def get_top20_art(data, ltm_key):
    arts = [{'artikelnr': a['artikelnr'], 'artikelnamn': a['artikelnamn'], 'value': a['ltm'].get(ltm_key, 0)} for a in data['articles'] if a['ltm'].get(ltm_key, 0) > 0]
    return sorted(arts, key=lambda x: x['value'], reverse=True)[:20]

def get_top20_cust(data, curr, prev):
    custs = []
    for c in data['customers']:
        cur_v = c['ltm'].get(curr, 0)
        if cur_v > 0:
            custs.append({'kund': c['kund'], 'current': cur_v, 'previous': c['ltm'].get(prev, 0), 'change': cur_v - c['ltm'].get(prev, 0)})
    return sorted(custs, key=lambda x: x['current'], reverse=True)[:20]


def generate_html(data, curr_ltm, prev_ltm):
    total_ltm = data['ltm_trend'].get(curr_ltm, 0)
    prev_ltm_val = data['ltm_trend'].get(prev_ltm, 0)
    yoy_chg = total_ltm - prev_ltm_val
    yoy_pct = (yoy_chg / prev_ltm_val * 100) if prev_ltm_val > 0 else 0
    
    cohorts = analyze_cohorts(data, curr_ltm, prev_ltm)
    churn_loss = sum(c['previous'] for c in cohorts['churned'])
    decline_loss = abs(sum(c['change'] for c in cohorts['declining']))
    growth_gain = sum(c['change'] for c in cohorts['growing'])
    new_gain = sum(c['current'] for c in cohorts['new'])
    active = len([c for c in data['customers'] if c['ltm'].get(curr_ltm, 0) > 0])
    
    top20_cust = get_top20_cust(data, curr_ltm, prev_ltm)
    top20_art = get_top20_art(data, curr_ltm)
    conc_pct = (sum(c['current'] for c in top20_cust) / total_ltm * 100) if total_ltm > 0 else 0
    
    months = sorted(data['monthly_totals'].keys())[-24:]
    m_labels = [datetime.strptime(m, '%Y-%m').strftime('%b %y') for m in months]
    m_values = [data['monthly_totals'].get(m, 0) for m in months]
    
    ltm_keys = sorted(data['ltm_trend'].keys())[-24:]
    l_labels, l_values = [], []
    month_map = {'jan': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr', 'maj': 'May', 'jun': 'Jun', 
                 'jul': 'Jul', 'aug': 'Aug', 'sep': 'Sep', 'okt': 'Oct', 'nov': 'Nov', 'dec': 'Dec'}
    for lk in ltm_keys:
        try:
            parts = lk.replace('LTM ', '').split('-')
            l_labels.append(f"{month_map.get(parts[1].lower(), parts[1])} {parts[0]}")
            l_values.append(data['ltm_trend'][lk])
        except: pass
    
    yoy_data = defaultdict(dict)
    for m in data['monthly_totals'].keys():
        try:
            dt = datetime.strptime(m, '%Y-%m')
            yoy_data[dt.month][dt.year] = data['monthly_totals'][m]
        except: pass
    years = sorted(set(y for md in yoy_data.values() for y in md.keys()))[-3:]
    yoy_ds = []
    colors = ['#64748B', '#4A9BA8', '#6366F1']
    for i, yr in enumerate(years):
        yoy_ds.append({'label': str(yr), 'data': [yoy_data[m].get(yr, 0) for m in range(1, 13)], 'backgroundColor': colors[i] if i < len(colors) else '#6366F1'})
    
    max_val = max(total_ltm, prev_ltm_val)
    scale = 160 / max_val if max_val > 0 else 1
    
    def cohort_rows(items, cols, limit=5):
        rows = []
        for c in items[:limit]:
            if cols == 2:
                rows.append(f'<tr><td class="customer-name">{c["kund"]}</td><td class="{"positive" if c.get("current",0) > 0 else "negative"}">{fmt_num(c.get("current", c.get("previous", 0)))}</td></tr>')
            else:
                cls = "positive" if c['change'] >= 0 else "negative"
                rows.append(f'<tr><td class="customer-name">{c["kund"]}</td><td>{fmt_num(c["current"])}</td><td>{fmt_num(c["previous"])}</td><td class="{cls}">{fmt_num(c["change"])}</td></tr>')
        return '\n'.join(rows)
    
    def top_cust_rows(items):
        rows = []
        for i, c in enumerate(items[:10]):
            cls = "positive" if c['change'] >= 0 else "negative"
            rows.append(f'<tr><td class="rank-cell">{i+1}</td><td>{c["kund"]}</td><td>{fmt_num(c["current"])}</td><td class="{cls}">{("+" if c["change"]>=0 else "")}{fmt_num(c["change"])}</td><td>{c["current"]/total_ltm*100:.1f}%</td></tr>')
        return '\n'.join(rows)
    
    def top_art_rows(items):
        rows = []
        for i, a in enumerate(items[:10]):
            rows.append(f'<tr><td class="rank-cell">{i+1}</td><td>{a["artikelnr"]}</td><td>{(a["artikelnamn"] or "")[:40]}</td><td>{fmt_num(a["value"])}</td><td>{a["value"]/total_ltm*100:.1f}%</td></tr>')
        return '\n'.join(rows)
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>HYAB Sales Intelligence Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
:root {{--bg-primary:#0F172A;--bg-card:#1E293B;--bg-card-hover:#334155;--text-primary:#F8FAFC;--text-secondary:#94A3B8;--text-muted:#64748B;--border:#334155;--red-churned:#DC2626;--red-churned-bg:rgba(220,38,38,0.15);--orange-declining:#F97316;--orange-declining-bg:rgba(249,115,22,0.15);--green-growing:#16A34A;--green-growing-bg:rgba(22,163,74,0.15);--teal-new:#0EA5E9;--teal-new-bg:rgba(14,165,233,0.15);--indigo-top:#6366F1;--indigo-top-bg:rgba(99,102,241,0.15);--positive:#22C55E;--negative:#EF4444;}}
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:'Inter',-apple-system,BlinkMacSystemFont,sans-serif;background:var(--bg-primary);color:var(--text-primary);line-height:1.5;}}
.dashboard{{max-width:1400px;margin:0 auto;padding:2rem;}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:2rem;padding-bottom:1.5rem;border-bottom:1px solid var(--border);}}
.header-left h1{{font-size:1.75rem;font-weight:700;letter-spacing:-0.025em;margin-bottom:0.25rem;}}
.header-left .period{{color:var(--text-secondary);font-size:0.9rem;}}
.header-right{{text-align:right;color:var(--text-muted);font-size:0.8rem;}}
.summary-strip{{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;margin-bottom:2rem;}}
.summary-card{{background:var(--bg-card);border-radius:12px;padding:1.25rem;border:1px solid var(--border);}}
.summary-card .label{{font-size:0.75rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:0.5rem;}}
.summary-card .value{{font-size:1.5rem;font-weight:700;font-family:'JetBrains Mono',monospace;}}
.summary-card .subvalue{{font-size:0.85rem;color:var(--text-secondary);margin-top:0.25rem;}}
.summary-card.highlight{{background:linear-gradient(135deg,var(--bg-card) 0%,rgba(99,102,241,0.1) 100%);border-color:var(--indigo-top);}}
.positive{{color:var(--positive);}} .negative{{color:var(--negative);}}
.chart-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:1.5rem;margin-bottom:2rem;}}
.chart-container{{background:var(--bg-card);border-radius:12px;padding:1.5rem;border:1px solid var(--border);}}
.chart-container h3{{font-size:0.85rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:1rem;}}
.chart-section{{background:var(--bg-card);border-radius:16px;padding:1.5rem;margin-bottom:2rem;border:1px solid var(--border);}}
.chart-section h2{{font-size:1rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:1rem;}}
.revenue-bridge{{background:var(--bg-card);border-radius:16px;padding:2rem;margin-bottom:2rem;border:1px solid var(--border);}}
.revenue-bridge h2{{font-size:1rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:1.5rem;}}
.bridge-container{{display:flex;align-items:flex-end;justify-content:space-between;height:200px;padding:0 1rem;}}
.bridge-item{{display:flex;flex-direction:column;align-items:center;flex:1;max-width:140px;}}
.bridge-bar{{width:60px;border-radius:4px 4px 0 0;}}
.bridge-bar.start{{background:var(--indigo-top);}} .bridge-bar.churned{{background:var(--red-churned);}}
.bridge-bar.declining{{background:var(--orange-declining);}} .bridge-bar.growing{{background:var(--green-growing);}}
.bridge-bar.new{{background:var(--teal-new);}} .bridge-bar.end{{background:linear-gradient(180deg,var(--indigo-top) 0%,#818CF8 100%);}}
.bridge-label{{margin-top:0.75rem;font-size:0.7rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);text-align:center;}}
.bridge-value{{margin-top:0.25rem;font-size:0.8rem;font-weight:600;font-family:'JetBrains Mono',monospace;}}
.bridge-connector{{flex:0.3;display:flex;align-items:center;justify-content:center;color:var(--text-muted);font-size:1.25rem;}}
.cohort-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:1.5rem;margin-bottom:2rem;}}
.cohort-card{{background:var(--bg-card);border-radius:12px;border:1px solid var(--border);overflow:hidden;}}
.cohort-header{{padding:1.25rem;display:flex;justify-content:space-between;align-items:center;}}
.cohort-header.churned{{background:var(--red-churned-bg);border-bottom:2px solid var(--red-churned);}}
.cohort-header.declining{{background:var(--orange-declining-bg);border-bottom:2px solid var(--orange-declining);}}
.cohort-header.growing{{background:var(--green-growing-bg);border-bottom:2px solid var(--green-growing);}}
.cohort-header.new{{background:var(--teal-new-bg);border-bottom:2px solid var(--teal-new);}}
.cohort-title{{font-size:0.75rem;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;}}
.cohort-header.churned .cohort-title{{color:var(--red-churned);}}
.cohort-header.declining .cohort-title{{color:var(--orange-declining);}}
.cohort-header.growing .cohort-title{{color:var(--green-growing);}}
.cohort-header.new .cohort-title{{color:var(--teal-new);}}
.cohort-count{{font-size:0.7rem;color:var(--text-secondary);margin-top:0.25rem;}}
.cohort-total{{font-family:'JetBrains Mono',monospace;font-size:1rem;font-weight:600;}}
.cohort-table{{width:100%;border-collapse:collapse;}}
.cohort-table th{{text-align:left;padding:0.75rem 1rem;font-size:0.65rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-muted);border-bottom:1px solid var(--border);font-weight:500;}}
.cohort-table th:last-child{{text-align:right;}}
.cohort-table td{{padding:0.6rem 1rem;font-size:0.8rem;border-bottom:1px solid var(--border);}}
.cohort-table td:last-child{{text-align:right;font-family:'JetBrains Mono',monospace;font-size:0.75rem;}}
.cohort-table tr:hover{{background:var(--bg-card-hover);}}
.customer-name{{color:var(--text-primary);font-weight:500;}}
.top-section{{background:var(--bg-card);border-radius:16px;border:1px solid var(--border);overflow:hidden;margin-bottom:2rem;}}
.top-header{{padding:1.25rem;background:var(--indigo-top-bg);border-bottom:2px solid var(--indigo-top);display:flex;justify-content:space-between;align-items:center;}}
.top-header h3{{font-size:0.75rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--indigo-top);font-weight:600;}}
.concentration-badge{{background:var(--indigo-top);color:white;padding:0.4rem 1rem;border-radius:20px;font-size:0.8rem;font-weight:600;}}
.top-table{{width:100%;border-collapse:collapse;}}
.top-table th{{text-align:left;padding:0.75rem 1rem;font-size:0.65rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-muted);border-bottom:1px solid var(--border);font-weight:500;}}
.top-table td{{padding:0.6rem 1rem;font-size:0.8rem;border-bottom:1px solid var(--border);}}
.top-table tr:hover{{background:var(--bg-card-hover);}}
.rank-cell{{font-weight:600;color:var(--text-muted);width:40px;}}
.footer{{text-align:center;padding-top:1.5rem;border-top:1px solid var(--border);color:var(--text-muted);font-size:0.75rem;}}
@media(max-width:1024px){{.summary-strip,.chart-grid,.cohort-grid{{grid-template-columns:1fr;}}}}
</style>
</head>
<body>
<div class="dashboard">
<header class="header"><div class="header-left"><h1>HYAB Sales Intelligence</h1><div class="period">{curr_ltm} vs {prev_ltm}</div></div><div class="header-right">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}<br>Data source: Sales Master File</div></header>

<div class="summary-strip">
<div class="summary-card highlight"><div class="label">Total LTM Revenue</div><div class="value">{fmt_sek(total_ltm)} SEK</div><div class="subvalue">{fmt_num(total_ltm)} SEK</div></div>
<div class="summary-card"><div class="label">Year-over-Year</div><div class="value {'positive' if yoy_pct >= 0 else 'negative'}">{yoy_pct:+.1f}%</div><div class="subvalue">{'+' if yoy_chg >= 0 else ''}{fmt_sek(yoy_chg)} SEK</div></div>
<div class="summary-card"><div class="label">Active Customers</div><div class="value">{active}</div><div class="subvalue">{len(cohorts['churned'])} churned · {len(cohorts['new'])} new</div></div>
<div class="summary-card"><div class="label">Movement</div><div class="value">{len(cohorts['growing'])} ↑ / {len(cohorts['declining'])} ↓</div><div class="subvalue">Growing vs declining</div></div>
</div>

<div class="chart-grid">
<div class="chart-container"><h3>📈 Rolling LTM Trend</h3><canvas id="ltmChart" height="200"></canvas></div>
<div class="chart-container"><h3>📊 Monthly Sales</h3><canvas id="monthlyChart" height="200"></canvas></div>
</div>

<div class="chart-section"><h2>📅 Year-over-Year Comparison by Month</h2><canvas id="yoyChart" height="120"></canvas></div>

<div class="revenue-bridge"><h2>Revenue Bridge</h2>
<div class="bridge-container">
<div class="bridge-item"><div class="bridge-bar start" style="height:{prev_ltm_val*scale:.0f}px;"></div><div class="bridge-label">Prior Year</div><div class="bridge-value">{fmt_sek(prev_ltm_val)}</div></div>
<div class="bridge-connector">→</div>
<div class="bridge-item"><div class="bridge-bar churned" style="height:{max(churn_loss*scale*0.1,10):.0f}px;"></div><div class="bridge-label">Churn Loss</div><div class="bridge-value negative">-{fmt_sek(churn_loss)}</div></div>
<div class="bridge-item"><div class="bridge-bar declining" style="height:{max(decline_loss*scale*0.1,10):.0f}px;"></div><div class="bridge-label">Decline Loss</div><div class="bridge-value negative">-{fmt_sek(decline_loss)}</div></div>
<div class="bridge-connector">+</div>
<div class="bridge-item"><div class="bridge-bar growing" style="height:{max(growth_gain*scale*0.1,10):.0f}px;"></div><div class="bridge-label">Growth Gain</div><div class="bridge-value positive">+{fmt_sek(growth_gain)}</div></div>
<div class="bridge-item"><div class="bridge-bar new" style="height:{max(new_gain*scale*0.1,10):.0f}px;"></div><div class="bridge-label">New Customers</div><div class="bridge-value positive">+{fmt_sek(new_gain)}</div></div>
<div class="bridge-connector">=</div>
<div class="bridge-item"><div class="bridge-bar end" style="height:{total_ltm*scale:.0f}px;"></div><div class="bridge-label">Current Year</div><div class="bridge-value">{fmt_sek(total_ltm)}</div></div>
</div>
</div>

<div class="cohort-grid">
<div class="cohort-card"><div class="cohort-header churned"><div><div class="cohort-title">⚠️ Churned Customers</div><div class="cohort-count">{len(cohorts['churned'])} accounts lost</div></div><div class="cohort-total negative">-{fmt_sek(churn_loss)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>Lost Revenue</th></tr></thead><tbody>{cohort_rows([{{'kund':c['kund'],'current':-c['previous']}} for c in cohorts['churned']],2)}</tbody></table></div>
<div class="cohort-card"><div class="cohort-header declining"><div><div class="cohort-title">📉 Declining Customers</div><div class="cohort-count">{len(cohorts['declining'])} accounts declining</div></div><div class="cohort-total negative">-{fmt_sek(decline_loss)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>Now</th><th>Was</th><th>Change</th></tr></thead><tbody>{cohort_rows(cohorts['declining'],4)}</tbody></table></div>
<div class="cohort-card"><div class="cohort-header growing"><div><div class="cohort-title">🎉 Growing Customers</div><div class="cohort-count">{len(cohorts['growing'])} accounts expanding</div></div><div class="cohort-total positive">+{fmt_sek(growth_gain)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>Now</th><th>Was</th><th>Growth</th></tr></thead><tbody>{cohort_rows(cohorts['growing'],4,7)}</tbody></table></div>
<div class="cohort-card"><div class="cohort-header new"><div><div class="cohort-title">✨ New Customers</div><div class="cohort-count">{len(cohorts['new'])} new accounts</div></div><div class="cohort-total positive">+{fmt_sek(new_gain)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>Revenue (LTM)</th></tr></thead><tbody>{cohort_rows(cohorts['new'],2,7)}</tbody></table></div>
</div>

<div class="top-section"><div class="top-header"><h3>👑 Top 20 Customers</h3><span class="concentration-badge">{conc_pct:.1f}% of Revenue</span></div><table class="top-table"><thead><tr><th>#</th><th>Customer</th><th>LTM Revenue</th><th>YoY Change</th><th>% of Total</th></tr></thead><tbody>{top_cust_rows(top20_cust)}</tbody></table></div>

<div class="top-section"><div class="top-header" style="background:rgba(74,155,168,0.15);border-color:#4A9BA8;"><h3 style="color:#4A9BA8;">📦 Top 20 Articles</h3><span class="concentration-badge" style="background:#4A9BA8;">{sum(a['value'] for a in top20_art)/total_ltm*100:.1f}% of Revenue</span></div><table class="top-table"><thead><tr><th>#</th><th>Article No</th><th>Description</th><th>LTM Revenue</th><th>% of Total</th></tr></thead><tbody>{top_art_rows(top20_art)}</tbody></table></div>

<footer class="footer"><p>HYAB Sales Intelligence Dashboard · Generated by HYAB Data App v3.0 · {datetime.now().strftime('%Y-%m-%d')}</p></footer>
</div>

<script>
Chart.defaults.color='#94A3B8';Chart.defaults.borderColor='#334155';
new Chart(document.getElementById('ltmChart'),{{type:'line',data:{{labels:{json.dumps(l_labels)},datasets:[{{label:'LTM Sales',data:{json.dumps(l_values)},borderColor:'#6366F1',backgroundColor:'rgba(99,102,241,0.1)',tension:0.3,fill:true,pointRadius:3}}]}},options:{{responsive:true,plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:false,ticks:{{callback:function(v){{return(v/1000000).toFixed(1)+'M';}}}}}}}}}});
new Chart(document.getElementById('monthlyChart'),{{type:'bar',data:{{labels:{json.dumps(m_labels)},datasets:[{{label:'Monthly Sales',data:{json.dumps(m_values)},backgroundColor:'#4A9BA8',borderRadius:4}}]}},options:{{responsive:true,plugins:{{legend:{{display:false}}}},scales:{{y:{{beginAtZero:true,ticks:{{callback:function(v){{return(v/1000000).toFixed(1)+'M';}}}}}}}}}});
new Chart(document.getElementById('yoyChart'),{{type:'bar',data:{{labels:['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'],datasets:{json.dumps(yoy_ds)}}},options:{{responsive:true,plugins:{{legend:{{position:'top'}}}},scales:{{y:{{beginAtZero:true,ticks:{{callback:function(v){{return(v/1000000).toFixed(1)+'M';}}}}}}}}}});
</script>
</body></html>'''
    return html


# =============================================================================
# MAIN APP
# =============================================================================

_page_title("Clean your data")
_section_label("Mode")
mode = st.radio("Mode", ["Order Book", "Sales", "Intelligence"], label_visibility="collapsed")
st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

if mode == "Order Book":
    col1, col2 = st.columns([1, 1])
    with col1:
        _section_label("Exchange rates")
        c1, c2, c3 = st.columns(3)
        with c1: eur = st.number_input("EUR", value=DEFAULT_FX['EUR'], step=0.01, format="%.2f")
        with c2: usd = st.number_input("USD", value=DEFAULT_FX['USD'], step=0.01, format="%.2f")
        with c3: gbp = st.number_input("GBP", value=DEFAULT_FX['GBP'], step=0.01, format="%.2f")
        fx = {'SEK': 1.0, 'EUR': eur, 'USD': usd, 'GBP': gbp}
        _section_label("Files")
        f = st.file_uploader("This week", type=['xlsx'], key="ob")
        if st.button("Process", type="primary", disabled=not f, use_container_width=True):
            if f:
                try:
                    wb = openpyxl.load_workbook(f)
                    ws = find_sheet(wb, ['Order book', 'Sheet1', 'Orders'])
                    if ws is None: st.error("Could not find order book sheet")
                    else:
                        orders = []
                        for r in range(2, ws.max_row + 1):
                            amt, cur = clean_amount(ws.cell(r, 6).value)
                            if amt is None: continue
                            od = ws.cell(r, 2).value
                            orders.append({'ordernr': ws.cell(r, 1).value, 'orderdatum': od if isinstance(od, datetime) else None, 'kundnamn': ws.cell(r, 3).value, 'belopp_sek': round(amt * fx.get(cur, 1.0), 2)})
                        st.session_state['ob_res'] = orders
                        st.rerun()
                except Exception as e: st.error(str(e))
    with col2:
        if 'ob_res' in st.session_state:
            _success_banner()
            orders = st.session_state['ob_res']
            c1, c2 = st.columns(2)
            with c1: st.metric("Orders", len(orders))
            with c2: st.metric("Total SEK", f"{sum(o['belopp_sek'] for o in orders):,.0f}")

elif mode == "Sales":
    col1, col2 = st.columns([1, 1])
    with col1:
        _section_label("Files")
        f = st.file_uploader("This month raw data", type=['xlsx'], key="sales")
        if st.button("Process", type="primary", disabled=not f, use_container_width=True):
            if f:
                try:
                    wb = openpyxl.load_workbook(f)
                    ws = find_sheet(wb, ['Article', 'Articles', 'Artikel'])
                    if ws is None: st.error("Could not find Article sheet")
                    else:
                        arts = []
                        for r in range(2, ws.max_row + 1):
                            an = ws.cell(r, 1).value
                            if an is None: continue
                            s = clean_num(ws.cell(r, 3).value)
                            if s is None or s == 0: continue
                            arts.append({'artikelnr': str(an), 'artikelnamn': ws.cell(r, 2).value, 'summa': s})
                        st.session_state['sales_res'] = sorted(arts, key=lambda x: x['summa'], reverse=True)
                        st.rerun()
                except Exception as e: st.error(str(e))
    with col2:
        if 'sales_res' in st.session_state:
            _success_banner()
            arts = st.session_state['sales_res']
            total = sum(a['summa'] for a in arts)
            c1, c2 = st.columns(2)
            with c1: st.metric("Total", f"{total:,.0f} SEK")
            with c2: st.metric("Articles", len(arts))

else:
    st.markdown('<div style="background:#E8F4FD;border-left:3px solid #1F4E79;padding:12px 16px;margin-bottom:16px;"><strong>Sales Intelligence Dashboard</strong><br><span style="font-size:13px;color:#666;">Upload Sales_work_file.xlsx to generate a downloadable HTML dashboard with all charts.</span></div>', unsafe_allow_html=True)
    f = st.file_uploader("Sales_work_file.xlsx", type=['xlsx'], key="intel")
    if st.button("Generate Dashboard", type="primary", disabled=not f, use_container_width=True):
        if f:
            try:
                with st.spinner("Loading master file..."):
                    wb = openpyxl.load_workbook(f, data_only=True)
                    data = parse_master(wb)
                ltms = sorted(data['ltm_trend'].keys())
                curr = ltms[-1] if ltms else None
                prev = ltms[-13] if len(ltms) > 12 else ltms[0] if ltms else None
                if curr and prev:
                    with st.spinner("Generating HTML dashboard..."):
                        html = generate_html(data, curr, prev)
                    st.session_state['intel_html'] = html
                    st.session_state['intel_data'] = data
                    st.session_state['intel_ltm'] = (curr, prev)
                    st.rerun()
                else: st.error("Could not find LTM data in file")
            except Exception as e:
                st.error(str(e))
                import traceback
                st.code(traceback.format_exc())
    
    if 'intel_html' in st.session_state:
        data = st.session_state['intel_data']
        curr, prev = st.session_state['intel_ltm']
        _success_banner()
        total = data['ltm_trend'].get(curr, 0)
        prev_v = data['ltm_trend'].get(prev, 0)
        yoy = ((total - prev_v) / prev_v * 100) if prev_v > 0 else 0
        c1, c2, c3, c4 = st.columns(4)
        with c1: st.metric("LTM Sales", f"{total/1e6:.1f}M SEK")
        with c2: st.metric("YoY Change", f"{yoy:+.1f}%")
        with c3: st.metric("Articles", f"{len(data['articles']):,}")
        with c4: st.metric("Customers", f"{len([c for c in data['customers'] if c['ltm'].get(curr,0)>0]):,}")
        st.markdown("---")
        st.download_button(label="📊 Download Dashboard (HTML)", data=st.session_state['intel_html'], file_name=f"HYAB_Dashboard_{datetime.now().strftime('%Y%m%d')}.html", mime="text/html", use_container_width=True)
        st.markdown('<div style="background:#F0FDF4;border-left:3px solid #16A34A;padding:12px 16px;margin-top:16px;"><strong>✓ Dashboard includes:</strong><br><span style="font-size:13px;color:#666;">• Rolling LTM trend chart<br>• Monthly sales bar chart<br>• Year-over-Year comparison by month<br>• Revenue bridge visualization<br>• Customer cohorts (Churned, Declining, Growing, New)<br>• Top 20 Customers with YoY change<br>• Top 20 Articles</span></div>', unsafe_allow_html=True)

st.markdown("---")
st.markdown('<div style="text-align:center;color:#6B7280;font-size:12px;">HYAB Data Cleaner v3.0 · Built for Victor</div>', unsafe_allow_html=True)
